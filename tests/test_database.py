import os
import sqlite3
import openpyxl
import pytest
import Projeto.database as database


@pytest.fixture(autouse=True)
def temp_db(monkeypatch, tmp_path):
    db_file = tmp_path / "test_database.db"
    excel_file = tmp_path / "test_devolucoes.xlsx"
    monkeypatch.setattr(database, "DB", str(db_file))
    monkeypatch.setattr(database, "EXCEL_FILE", str(excel_file))
    database.criar()
    yield
    # cleanup is automatic with tmp_path


def test_inserir_e_listar():
    dados = {
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-0001",
        "modelo": "Teste",
        "serial": "SN0001",
        "status": "Devolvido",
        "motivo": "Teste",
        "foto": None,
    }

    database.inserir(dados)
    resultados = database.listar()

    assert len(resultados) == 1
    registro = resultados[0]
    assert registro[3] == "Test User"
    assert registro[8] == "SN0001"


def test_exportar_para_excel():
    dados = {
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-0002",
        "modelo": "ModeloX",
        "serial": "SN0002",
        "status": "Danificado",
        "motivo": "Teste export",
        "foto": None,
    }
    database.inserir(dados)
    assert os.path.exists(database.EXCEL_FILE)

    wb = openpyxl.load_workbook(database.EXCEL_FILE)
    assert "Devolucoes" in wb.sheetnames
    ws = wb["Devolucoes"]
    # header + 1 registro
    assert ws.max_row == 2
    assert ws["A2"].value == 1
    assert ws["C2"].value == "test.user"  # usuario
