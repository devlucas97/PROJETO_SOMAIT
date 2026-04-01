import logging
import os
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from Projeto.database import criar, inserir, inserir_com_data, listar
import openpyxl

# Import com fallback para compatibilidade Linux/Windows
try:
    from Projeto.email_service import enviar_email, email_dano
    OUTLOOK_AVAILABLE = True
except ImportError:
    OUTLOOK_AVAILABLE = False
    enviar_email = None
    email_dano = None

# Configuração de logging
logger = logging.getLogger(__name__)
if not logger.hasHandlers():
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s: %(message)s"))
    logger.addHandler(handler)
logger.setLevel(logging.INFO)

# Criar app Flask
app = Flask(__name__)
app.secret_key = "chave_secreta_dev"

# Criar diretórios necessários
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Inicializar banco de dados
criar()

@app.route("/")
def index():
    """Página inicial com listagem de devoluções"""
    try:
        dados = listar()
        search_id = request.args.get("search_id", "").strip()

        if search_id:
            if not search_id.isdigit():
                flash("ID deve ser um número inteiro", "error")
                dados_filtrados = []
            else:
                id_busca = int(search_id)
                dados_filtrados = [registro for registro in dados if registro[0] == id_busca]
                if not dados_filtrados:
                    flash(f"Nenhum registro encontrado para ID {id_busca}", "warning")
            dados = dados_filtrados
        
        logger.info(f"Listagem de {len(dados)} registros obtida")
        return render_template("index.html", dados=dados, search_id=search_id)
    except Exception as err:
        logger.exception("Erro ao listar devoluções")
        flash("Erro ao carregar dados", "error")
        return render_template("index.html", dados=[], search_id="")

@app.route("/nova", methods=["GET", "POST"])
def nova_devolucao():
    """Formulário para nova devolução"""
    if request.method == "POST":
        try:
            # Coletar dados do formulário
            dados = {
                "usuario": request.form.get("usuario"),
                "nome": request.form.get("nome"),
                "matricula": request.form.get("matricula"),
                "departamento": request.form.get("departamento"),
                "patrimonio": request.form.get("patrimonio"),
                "modelo": request.form.get("modelo"),
                "serial": request.form.get("serial"),
                "status": request.form.get("status"),
                "motivo": request.form.get("motivo"),
                "foto": None
            }

            # Validação básica
            campos_obrigatorios = ["usuario", "nome", "matricula", "departamento", "patrimonio", "modelo", "serial"]
            if not all(dados.get(campo) for campo in campos_obrigatorios):
                flash("Todos os campos obrigatórios devem ser preenchidos", "error")
                return redirect(url_for("nova_devolucao"))

            # Processar upload de foto (se houver e status for Danificado)
            if "foto" in request.files and dados["status"] == "Danificado":
                arquivo = request.files["foto"]
                if arquivo and arquivo.filename:
                    try:
                        ext = os.path.splitext(arquivo.filename)[1]
                        nome_arquivo = f"{dados['patrimonio']}{ext}"
                        destino = os.path.join(UPLOAD_DIR, nome_arquivo)
                        arquivo.save(destino)
                        dados["foto"] = destino
                        logger.info(f"Foto salva: {destino}")
                    except Exception as err:
                        logger.warning(f"Erro ao salvar foto: {err}")
                        dados["foto"] = None

            # Inserir no banco
            inserir(dados)
            logger.info(f"Devolução registrada: {dados['patrimonio']}")

            # Tentar enviar email (não bloqueia se falhar)
            if OUTLOOK_AVAILABLE:
                try:
                    if dados["status"] == "Danificado":
                        email_dano(dados)
                        logger.info(f"Email de dano enviado para {dados['nome']}")
                    else:
                        enviar_email(dados)
                        logger.info(f"Email de devolução enviado para {dados['nome']}")
                except RuntimeError as err:
                    logger.warning(f"Email não enviado: {err}")
                    flash("Devolução registrada, mas email não foi enviado (Outlook indisponível)", "warning")
            else:
                logger.info("Outlook não disponível nesta plataforma (Linux/Mac sem win32com)")
                flash("Devolução registrada (email indisponível nesta plataforma)", "info")

            return redirect(url_for("index"))

        except Exception as err:
            logger.exception("Erro ao processar nova devolução")
            flash("Erro ao processar a devolução", "error")
            return redirect(url_for("nova_devolucao"))

    return render_template("nova.html")

@app.route("/foto/<filename>")
def servir_foto(filename):
    """Serve a imagem do upload"""
    try:
        caminho_arquivo = os.path.join(UPLOAD_DIR, filename)
        if os.path.exists(caminho_arquivo):
            logger.info(f"Foto servida: {filename}")
            return send_file(caminho_arquivo, mimetype='image/png')
        else:
            logger.warning(f"Arquivo não encontrado: {filename}")
            flash("Foto não encontrada", "error")
            return redirect(url_for("index"))
    except Exception as err:
        logger.exception(f"Erro ao servir foto: {err}")
        flash("Erro ao carregar foto", "error")
        return redirect(url_for("index"))

@app.route("/visualizar/<filename>")
def visualizar_foto(filename):
    """Página para visualizar a foto"""
    try:
        caminho_arquivo = os.path.join(UPLOAD_DIR, filename)
        if os.path.exists(caminho_arquivo):
            logger.info(f"Visualizando foto: {filename}")
            return render_template("visualizar_foto.html", foto_url=url_for("servir_foto", filename=filename), nome_arquivo=filename)
        else:
            logger.warning(f"Arquivo não encontrado: {filename}")
            flash("Foto não encontrada", "error")
            return redirect(url_for("index"))
    except Exception as err:
        logger.exception(f"Erro ao visualizar foto: {err}")
        flash("Erro ao carregar foto", "error")
        return redirect(url_for("index"))

@app.route("/export-excel")
def exportar_excel():
    """Exporta os registros de devolução para planilha Excel"""
    try:
        registros = listar()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devolucoes"

        headers = ["ID", "Data", "Usuario", "Nome", "Matricula", "Departamento", "Patrimonio", "Modelo", "Serial", "Status", "Motivo", "Foto"]
        ws.append(headers)

        for reg in registros:
            ws.append(list(reg))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("Exportação para Excel gerada com sucesso")
        return send_file(output, download_name="devolucoes.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as err:
        logger.exception(f"Erro ao exportar para Excel: {err}")
        flash("Erro ao exportar para Excel", "error")
        return redirect(url_for("index"))

@app.route("/import-excel", methods=["GET", "POST"])
def importar_excel():
    """Importa registros de planilha Excel para o banco de dados"""
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        if not arquivo or arquivo.filename == "":
            flash("Nenhum arquivo selecionado", "error")
            return redirect(url_for("importar_excel"))

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active
            total = 0
            importados = 0

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue
                _, data, usuario, nome, matricula, departamento, patrimonio, modelo, serial, status, motivo, foto = row
                total += 1
                if not (usuario and nome and matricula and departamento and patrimonio and modelo and serial and status):
                    logger.warning(f"Linha {idx} omitida por falta de campos obrigatórios")
                    continue

                try:
                    inserir_com_data({
                        "data": data if data else None,
                        "usuario": usuario,
                        "nome": nome,
                        "matricula": matricula,
                        "departamento": departamento,
                        "patrimonio": patrimonio,
                        "modelo": modelo,
                        "serial": serial,
                        "status": status,
                        "motivo": motivo,
                        "foto": foto
                    })
                    importados += 1
                except Exception as ex:
                    logger.warning(f"Erro ao inserir linha {idx}: {ex}")

            flash(f"Importação finalizada: {importados}/{total} registros importados", "success")
            logger.info(f"Importação concluída: {importados}/{total}")
            return redirect(url_for("index"))
        except Exception as err:
            logger.exception(f"Erro ao importar Excel: {err}")
            flash("Erro ao importar planilha Excel", "error")
            return redirect(url_for("importar_excel"))

    return render_template("importar_excel.html")

if __name__ == "__main__":
    logger.info("Iniciando servidor Flask em http://127.0.0.1:5000")
    app.run(debug=True, host="127.0.0.1", port=5000)
