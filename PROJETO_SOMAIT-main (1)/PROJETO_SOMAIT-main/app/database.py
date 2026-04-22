import json
import sqlite3
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
import os
import re
import unicodedata
from collections import Counter

from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.logging_config import get_logger
from app.runtime_paths import get_runtime_path, iter_config_paths

logger = get_logger(__name__)

# ── Configuração ─────────────────────────────────────────────────
DB = get_runtime_path("database.db")
PLANILHA_EMPRESA = os.getenv("PLANILHA_EMPRESA", "")
ABA_PLANILHA_EMPRESA = os.getenv("ABA_PLANILHA_EMPRESA", "Devolucoes")
MAPEAMENTO_COLUNAS_PLANILHA = os.getenv("MAPEAMENTO_COLUNAS_PLANILHA", "")

# Status do workflow Dell — fonte única de verdade
DELL_WF_STATUSES = frozenset({
    "Aguardando Cotação", "Cotação Recebida",
    "Reparo Aprovado", "Em Reparo", "Concluído",
})


def _normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def _obter_mapeamento_colunas():
    mapeamento = {
        "id": "id",
        "data": "data",
        "usuario": "usuario",
        "nome": "nome",
        "matricula": "matricula",
        "departamento": "departamento",
        "patrimonio": "patrimonio",
        "modelo": "modelo",
        "serial": "serial",
        "status": "status",
        "motivo": "motivo",
        "foto": "foto",
        "colaborador": "nome",
        "nome_completo": "nome",
        "setor": "departamento",
        "area": "departamento",
        "patrimonio_ti": "patrimonio",
        "numero_serie": "serial",
        "serie": "serial",
        "situacao": "status",
        "observacao": "motivo",
    }

    bruto = MAPEAMENTO_COLUNAS_PLANILHA.strip()
    if not bruto:
        return mapeamento

    # Formato esperado: campo=Nome da Coluna;campo2=Outra Coluna
    for trecho in re.split(r"[;,]", bruto):
        item = trecho.strip()
        if not item or "=" not in item:
            continue
        campo, coluna = item.split("=", 1)
        campo_norm = _normalizar_texto(campo)
        coluna_norm = _normalizar_texto(coluna)
        if campo_norm and coluna_norm:
            mapeamento[coluna_norm] = campo_norm

    return mapeamento


def _row_to_dict(row):
    """Converte sqlite3.Row para dict padrão (compatível com Jinja2 e serialização)."""
    if row is None:
        return None
    return dict(row)


def _agora_brasilia():
    """Retorna data/hora atual em Brasília, com fallback para ambientes sem tzdata."""
    try:
        return datetime.now(ZoneInfo("America/Sao_Paulo"))
    except ZoneInfoNotFoundError:
        logger.warning(
            "Base de fuso horário ausente; usando UTC-03:00 fixo para America/Sao_Paulo."
        )
        return datetime.now(timezone(timedelta(hours=-3)))


def conectar():
    conn = sqlite3.connect(DB, detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    return conn
 
def criar():
    with conectar() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS devolucoes (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            data                TEXT,
            usuario             TEXT,
            nome                TEXT,
            matricula           TEXT,
            departamento        TEXT,
            patrimonio          TEXT,
            modelo              TEXT,
            serial              TEXT,
            status              TEXT,
            motivo              TEXT,
            foto                TEXT,
            diretoria           TEXT,
            tipo                TEXT,
            marca               TEXT,
            processador         TEXT,
            memoria             TEXT,
            armazenamento       TEXT,
            possui_carregador   TEXT,
            recebido_por        TEXT,
            unidade             TEXT,
            observacoes         TEXT,
            movido_para_estoque TEXT,
            email_responsavel   TEXT,
            email_enviado_em    TEXT,
            gestor_email        TEXT,
            chamado_dell        TEXT
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            login       TEXT UNIQUE NOT NULL,
            nome        TEXT NOT NULL,
            senha_hash  TEXT NOT NULL,
            admin       INTEGER NOT NULL DEFAULT 0,
            ativo       INTEGER NOT NULL DEFAULT 1,
            criado_em   TEXT
        )
        """)
        conn.commit()
    _migrar()


def _migrar():
    """Adiciona colunas novas e índices em bancos existentes (idempotente)."""
    novas = [
        "diretoria", "tipo", "marca", "processador", "memoria",
        "armazenamento", "possui_carregador", "recebido_por",
        "unidade", "observacoes", "movido_para_estoque",
        "email_responsavel", "email_enviado_em",
        "gestor_email", "chamado_dell", "deletado_em", "registrado_por",
    ]
    with conectar() as conn:
        existentes = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(devolucoes)").fetchall()
        }
        for col in novas:
            if col not in existentes:
                # Nomes de coluna são literals controlados, não input de usuário
                conn.execute(f"ALTER TABLE devolucoes ADD COLUMN {col} TEXT")
                logger.info("Coluna adicionada ao banco: %s", col)
        # Índices para acelerar buscas e filtros frequentes
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_devolucoes_status ON devolucoes(status)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_devolucoes_busca ON devolucoes("
            "nome, patrimonio, usuario, departamento, modelo, serial, marca)"
        )
        conn.commit()


def _obter_config_planilha():
    """Lê caminho da planilha do config.json (prioridade) ou variáveis de ambiente."""
    planilha = PLANILHA_EMPRESA
    aba = ABA_PLANILHA_EMPRESA
    # Só usa config.json quando as variáveis de módulo estão em branco
    if not planilha:
        for config_file in iter_config_paths():
            if not os.path.exists(config_file):
                continue
            try:
                with open(config_file, encoding="utf-8") as f:
                    cfg = json.load(f)
                planilha = cfg.get("planilha_empresa") or planilha
                aba = cfg.get("aba_planilha_empresa") or aba
                if planilha:
                    break
            except Exception as err:
                logger.warning("Falha ao ler config.json da planilha: %s", err)
    # Remove aspas extras que o usuário possa ter digitado ao redor do caminho
    planilha = planilha.strip().strip('"').strip("'").strip() if planilha else ""
    return planilha, aba


def _parece_caminho_windows(caminho):
    return bool(re.match(r"^[A-Za-z]:[\\/]", caminho or "")) or (caminho or "").startswith("\\\\")


def _descrever_ambiente_execucao():
    if os.name == "nt":
        return "Windows"
    if os.path.exists("/.dockerenv"):
        return "Linux / container"
    return "Linux"


def diagnosticar_config_planilha(planilha=None):
    caminho = planilha if planilha is not None else _obter_config_planilha()[0]
    caminho = caminho.strip().strip('"').strip("'").strip() if caminho else ""
    ambiente = _descrever_ambiente_execucao()

    if not caminho:
        return {
            "caminho": "",
            "ambiente": ambiente,
            "ok": True,
            "bloqueada": False,
            "nivel": "info",
            "mensagem": "Sincronização automática desabilitada.",
            "detalhe": "Informe um caminho de planilha para ativar a integração Excel.",
        }

    if os.name != "nt" and _parece_caminho_windows(caminho):
        return {
            "caminho": caminho,
            "ambiente": ambiente,
            "ok": False,
            "bloqueada": True,
            "nivel": "error",
            "mensagem": "Caminho Windows detectado em ambiente Linux / container.",
            "detalhe": "Para evitar criar um arquivo local com nome literal, a sincronização Excel fica bloqueada. Use um caminho acessível dentro do container, como /workspaces/..., ou execute a aplicação no Windows host.",
        }

    return {
        "caminho": caminho,
        "ambiente": ambiente,
        "ok": True,
        "bloqueada": False,
        "nivel": "success",
        "mensagem": "Caminho compatível com o ambiente atual.",
        "detalhe": "A integração Excel pode usar este caminho para sincronização automática.",
    }


def _descrever_erro_sincronizacao(err):
    planilha, _ = _obter_config_planilha()
    if isinstance(err, PermissionError):
        return (
            "A devolução foi salva, mas a planilha Excel está bloqueada para escrita. "
            f"Feche o arquivo '{os.path.basename(planilha)}' no Excel e sincronize novamente."
        )
    return f"A devolução foi salva, mas a sincronização da planilha falhou: {err}"


def _sincronizar_planilha_seguro(contexto):
    try:
        sincronizar_planilha_completa()
        return None
    except Exception as err:
        logger.warning("Falha ao sincronizar planilha %s: %s", contexto, err)
        return _descrever_erro_sincronizacao(err)


# ── Planilha corporativa — estilos e layout compartilhados ───────

_FILL_HDR  = PatternFill("solid", fgColor="0F172A")
_FILL_COL  = PatternFill("solid", fgColor="1E40AF")
_FILL_OK   = PatternFill("solid", fgColor="D1FAE5")
_FILL_DANO = PatternFill("solid", fgColor="FEE2E2")
_FILL_PEND = PatternFill("solid", fgColor="FEF3C7")
_FILL_DELL = PatternFill("solid", fgColor="DBEAFE")
_FILL_ALT  = PatternFill("solid", fgColor="F8FAFC")

EXCEL_HEADERS = [
    "DATA", "SETOR", "DIRETORIA", "LOGIN DE REDE",
    "TIPO", "MARCA", "MODELO", "PROCESSADOR",
    "MEMÓRIA", "ARMAZENAMENTO", "TAG", "POSSUI CARREGADOR",
    "ENTREGUE POR", "RECEBIDO POR", "MOTIVO",
    "UNIDADE", "OBSERVAÇÕES", "MOV. ESTOQUE",
    "SITUAÇÃO", "Nº CHAMADO DELL", "GESTOR EMAIL",
]

EXCEL_WIDTHS = [
    16, 16, 18, 18, 14, 14, 24, 18, 12, 16, 14, 16,
    22, 22, 28, 14, 28, 18, 18, 20, 28,
]


def row_fill(status):
    """Retorna o fill do Excel adequado ao status (reutilizado por web.py na exportação)."""
    if status == "OK":
        return _FILL_OK
    if status == "Danificado":
        return _FILL_DANO
    if status == "Pendente":
        return _FILL_PEND
    if status in DELL_WF_STATUSES:
        return _FILL_DELL
    return None


def registro_para_linha_excel(r):
    """Converte um registro (dict) na ordem das colunas da planilha Excel."""
    return [
        r.get("data"), r.get("departamento"), r.get("diretoria"), r.get("usuario"),
        r.get("tipo"), r.get("marca"), r.get("modelo"), r.get("processador"),
        r.get("memoria"), r.get("armazenamento"), r.get("patrimonio"), r.get("possui_carregador"),
        r.get("nome"), r.get("recebido_por"), r.get("motivo"), r.get("unidade"),
        r.get("observacoes"), r.get("movido_para_estoque"), r.get("status"),
        r.get("chamado_dell"), r.get("gestor_email"),
    ]


def _escrever_aba_excel(ws, titulo, fill_titulo, registros):
    """Limpa e reescreve uma aba com título, cabeçalho e dados (reutilizável)."""
    n_cols = len(EXCEL_HEADERS)

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill = _FILL_COL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    for row_idx, reg in enumerate(registros, 3):
        status = reg.get("status") or ""
        fill = row_fill(status) or (_FILL_ALT if row_idx % 2 == 0 else None)
        for col_idx, value in enumerate(registro_para_linha_excel(reg), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    for i, w in enumerate(EXCEL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def _obter_ou_criar_aba(wb, nome, posicao=None):
    """Retorna uma aba existente ou cria uma nova."""
    if nome not in wb.sheetnames:
        return wb.create_sheet(nome, posicao)
    return wb[nome]


# ── Planilha — alimentação e sincronização ───────────────────────

def alimentar_planilha_existente(registro):
    """Atualiza uma planilha corporativa existente sem criar novos arquivos."""
    planilha, aba_nome = _obter_config_planilha()
    if not planilha:
        return

    diagnostico = diagnosticar_config_planilha(planilha)
    if diagnostico["bloqueada"]:
        logger.warning(
            "Integração Excel bloqueada: %s | caminho=%s | ambiente=%s",
            diagnostico["mensagem"], planilha, diagnostico["ambiente"],
        )
        return

    if not os.path.exists(planilha):
        logger.warning("Planilha corporativa não encontrada: %s", planilha)
        return

    wb = openpyxl.load_workbook(planilha)
    if aba_nome not in wb.sheetnames:
        logger.warning("Aba '%s' não encontrada na planilha corporativa", aba_nome)
        return

    ws = wb[aba_nome]
    headers = [_normalizar_texto(c.value) for c in ws[1]]
    possui_cabecalho = any(headers)
    mapeamento = _obter_mapeamento_colunas()

    dados_registro = dict(registro) if not isinstance(registro, dict) else registro

    if possui_cabecalho:
        linha = [dados_registro.get(mapeamento.get(h, "")) for h in headers]
    else:
        linha = [dados_registro.get(col) for col in dados_registro]

    ws.append(linha)
    wb.save(planilha)
    logger.info("Planilha corporativa alimentada com registro ID %s", dados_registro.get("id"))


def sincronizar_planilha_completa():
    """Reescreve TODOS os registros em 3 abas fixas na planilha corporativa.

    Cria o arquivo automaticamente se não existir. As abas são:
    - "Devoluções"       — todos os registros com cores por situação
    - "Resumo por Status"— contagem por situação
    - "Dell Danificados" — apenas Dell com dano ou workflow de reparo
    """
    planilha, _ = _obter_config_planilha()
    if not planilha:
        return

    diagnostico = diagnosticar_config_planilha(planilha)
    if diagnostico["bloqueada"]:
        logger.warning(
            "Sincronização Excel bloqueada: %s | caminho=%s | ambiente=%s",
            diagnostico["mensagem"], planilha, diagnostico["ambiente"],
        )
        return

    diretorio = os.path.dirname(planilha)
    if diretorio:
        os.makedirs(diretorio, exist_ok=True)

    if os.path.exists(planilha):
        wb = openpyxl.load_workbook(planilha)
    else:
        wb = openpyxl.Workbook()
        for nome in wb.sheetnames:
            del wb[nome]
        logger.info("Planilha corporativa criada automaticamente: %s", planilha)

    registros = listar()

    # ── ABA 1: Devoluções ─────────────────────────────────────
    ws1 = _obter_ou_criar_aba(wb, "Devoluções", 0)
    _escrever_aba_excel(
        ws1,
        f"AZZAS TI — Gestão de Devoluções ({len(registros)} registros)",
        _FILL_HDR,
        registros,
    )

    # ── ABA 2: Resumo por Status ──────────────────────────────
    ws2 = _obter_ou_criar_aba(wb, "Resumo por Status")
    if ws2.max_row:
        ws2.delete_rows(1, ws2.max_row)

    ws2["A1"] = "Situação"
    ws2["B1"] = "Quantidade"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = _FILL_HDR
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    contagem = Counter((r["status"] or "Sem status") for r in registros)
    all_statuses = [
        "OK", "Danificado", "Pendente",
        "Aguardando Cotação", "Cotação Recebida",
        "Reparo Aprovado", "Em Reparo", "Concluído",
    ]
    row2 = 2
    total2 = 0
    for s in all_statuses:
        if s in contagem:
            c = ws2.cell(row=row2, column=1, value=s)
            v = ws2.cell(row=row2, column=2, value=contagem[s])
            f = row_fill(s)
            if f:
                c.fill = f
                v.fill = f
            c.font = Font(name="Calibri", size=11)
            v.font = Font(name="Calibri", size=11)
            v.alignment = Alignment(horizontal="center")
            total2 += contagem[s]
            row2 += 1
    for s, cnt in contagem.items():
        if s not in all_statuses:
            ws2.cell(row=row2, column=1, value=s).font = Font(name="Calibri", size=11)
            c2 = ws2.cell(row=row2, column=2, value=cnt)
            c2.font = Font(name="Calibri", size=11)
            c2.alignment = Alignment(horizontal="center")
            total2 += cnt
            row2 += 1
    t = ws2.cell(row=row2, column=1, value="TOTAL")
    tv = ws2.cell(row=row2, column=2, value=total2)
    t.font = Font(name="Calibri", bold=True, size=11)
    tv.font = Font(name="Calibri", bold=True, size=11)
    tv.alignment = Alignment(horizontal="center")

    # ── ABA 3: Dell Danificados ───────────────────────────────
    dell_statuses = {"Danificado"} | DELL_WF_STATUSES
    dell_rows = [
        r for r in registros
        if (r["marca"] or "").strip().lower() == "dell"
        and (r["status"] or "") in dell_statuses
    ]
    ws3 = _obter_ou_criar_aba(wb, "Dell Danificados")
    _escrever_aba_excel(
        ws3,
        f"AZZAS TI — Dell Danificados / Em Reparo ({len(dell_rows)} registros)",
        PatternFill("solid", fgColor="7F1D1D"),
        dell_rows,
    )

    wb.active = wb["Devoluções"]
    # Escrita atômica: grava em arquivo temporário e renomeia
    tmp_planilha = planilha + ".tmp"
    try:
        wb.save(tmp_planilha)
        os.replace(tmp_planilha, planilha)
    except Exception:
        # Remove temporário em caso de falha
        if os.path.exists(tmp_planilha):
            os.remove(tmp_planilha)
        raise
    logger.info(
        "Planilha corporativa sincronizada: %d registro(s), %d Dell — '%s'",
        len(registros), len(dell_rows), planilha,
    )
 
def inserir(dados):
    """Insere uma nova devolução e retorna o registro como dict."""
    with conectar() as conn:
        horario_brasilia = _agora_brasilia()
        cursor = conn.execute("""
        INSERT INTO devolucoes (
            data, usuario, nome, matricula, departamento,
            patrimonio, modelo, serial, status, motivo, foto,
            diretoria, tipo, marca, processador, memoria,
            armazenamento, possui_carregador, recebido_por,
            unidade, observacoes, movido_para_estoque,
            email_responsavel, email_enviado_em, gestor_email, chamado_dell,
            registrado_por
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            horario_brasilia.strftime("%d/%m/%Y %H:%M"),
            dados.get("usuario"), dados.get("nome"),
            dados.get("matricula"), dados.get("departamento"),
            dados.get("patrimonio"), dados.get("modelo"),
            dados.get("serial"), dados.get("status"),
            dados.get("motivo"), dados.get("foto"),
            dados.get("diretoria"), dados.get("tipo"),
            dados.get("marca"), dados.get("processador"),
            dados.get("memoria"), dados.get("armazenamento"),
            dados.get("possui_carregador"), dados.get("recebido_por"),
            dados.get("unidade"), dados.get("observacoes"),
            dados.get("movido_para_estoque"), dados.get("email_responsavel"),
            None,  # email_enviado_em
            dados.get("gestor_email"), dados.get("chamado_dell"),
            dados.get("registrado_por"),
        ))
        novo_id = cursor.lastrowid
        conn.commit()
        registro = conn.execute(
            "SELECT * FROM devolucoes WHERE id = ?", (novo_id,)
        ).fetchone()

    registro_dict = _row_to_dict(registro)
    sync_error = _sincronizar_planilha_seguro("após nova devolução")
    if registro_dict is not None:
        registro_dict["_sync_error"] = sync_error
    return registro_dict


def listar():
    """Retorna todas as devoluções ativas (não deletadas) como lista de dicts."""
    with conectar() as conn:
        rows = conn.execute(
            "SELECT * FROM devolucoes WHERE deletado_em IS NULL ORDER BY id DESC"
        ).fetchall()
    return [_row_to_dict(r) for r in rows]


def buscar_por_id(id_devolucao):
    """Retorna um registro ativo como dict ou None."""
    with conectar() as conn:
        row = conn.execute(
            "SELECT * FROM devolucoes WHERE id = ? AND deletado_em IS NULL", (id_devolucao,)
        ).fetchone()
    return _row_to_dict(row)


def estatisticas():
    """Retorna contagens por status para o dashboard executivo."""
    with conectar() as conn:
        total      = conn.execute("SELECT COUNT(*) FROM devolucoes WHERE deletado_em IS NULL").fetchone()[0]
        ok         = conn.execute("SELECT COUNT(*) FROM devolucoes WHERE status = ? AND deletado_em IS NULL", ("OK",)).fetchone()[0]
        danificado = conn.execute("SELECT COUNT(*) FROM devolucoes WHERE status = ? AND deletado_em IS NULL", ("Danificado",)).fetchone()[0]
        pendente   = conn.execute("SELECT COUNT(*) FROM devolucoes WHERE status = ? AND deletado_em IS NULL", ("Pendente",)).fetchone()[0]
    return {"total": total, "ok": ok, "danificado": danificado, "pendente": pendente}


def _construir_filtro(status=None, busca=None, data_inicio=None, data_fim=None):
    """Constrói a cláusula WHERE e parâmetros para filtros reutilizáveis."""
    where = "WHERE deletado_em IS NULL"
    params = []
    if status:
        where += " AND status = ?"
        params.append(status)
    if busca:
        like = f"%{busca}%"
        where += (
            " AND (nome LIKE ? OR matricula LIKE ? OR patrimonio LIKE ?"
            " OR modelo LIKE ? OR serial LIKE ? OR usuario LIKE ? OR departamento LIKE ?"
            " OR diretoria LIKE ? OR tipo LIKE ? OR marca LIKE ? OR recebido_por LIKE ?"
            " OR unidade LIKE ?)"
        )
        params.extend([like] * 12)
    if data_inicio:
        where += " AND substr(data,7,4)||'-'||substr(data,4,2)||'-'||substr(data,1,2) >= ?"
        params.append(data_inicio)
    if data_fim:
        where += " AND substr(data,7,4)||'-'||substr(data,4,2)||'-'||substr(data,1,2) <= ?"
        params.append(data_fim)
    return where, params


def contar_filtrado(status=None, busca=None, data_inicio=None, data_fim=None):
    """Retorna a contagem total de registros que atendem aos filtros."""
    where, params = _construir_filtro(status, busca, data_inicio, data_fim)
    with conectar() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM devolucoes {where}", params).fetchone()[0]
    return total


def listar_filtrado(status=None, busca=None, data_inicio=None, data_fim=None,
                    limit=None, offset=None):
    """Lista devoluções com filtros opcionais e paginação SQL. Retorna lista de dicts."""
    where, params = _construir_filtro(status, busca, data_inicio, data_fim)
    query = f"SELECT * FROM devolucoes {where} ORDER BY id DESC"
    if limit is not None:
        query += " LIMIT ?"
        params.append(limit)
        if offset is not None:
            query += " OFFSET ?"
            params.append(offset)
    with conectar() as conn:
        rows = conn.execute(query, params).fetchall()
    return [_row_to_dict(r) for r in rows]


def atualizar(id_devolucao, dados):
    """Atualiza os campos de um registro existente."""
    with conectar() as conn:
        # Se foto foi enviada, atualizar; caso contrário, manter a existente
        if dados.get("foto"):
            conn.execute("""
            UPDATE devolucoes SET
                usuario = ?, nome = ?, matricula = ?, departamento = ?,
                patrimonio = ?, modelo = ?, serial = ?, status = ?, motivo = ?,
                foto = ?,
                diretoria = ?, tipo = ?, marca = ?, processador = ?, memoria = ?,
                armazenamento = ?, possui_carregador = ?, recebido_por = ?,
                unidade = ?, observacoes = ?, movido_para_estoque = ?, email_responsavel = ?,
                gestor_email = ?, chamado_dell = ?
            WHERE id = ?
            """, (
                dados.get("usuario"), dados.get("nome"),
                dados.get("matricula"), dados.get("departamento"),
                dados.get("patrimonio"), dados.get("modelo"),
                dados.get("serial"), dados.get("status"),
                dados.get("motivo"), dados.get("foto"),
                dados.get("diretoria"), dados.get("tipo"),
                dados.get("marca"), dados.get("processador"),
                dados.get("memoria"), dados.get("armazenamento"),
                dados.get("possui_carregador"), dados.get("recebido_por"),
                dados.get("unidade"), dados.get("observacoes"),
                dados.get("movido_para_estoque"), dados.get("email_responsavel"),
                dados.get("gestor_email"), dados.get("chamado_dell"),
                id_devolucao,
            ))
        else:
            conn.execute("""
            UPDATE devolucoes SET
                usuario = ?, nome = ?, matricula = ?, departamento = ?,
                patrimonio = ?, modelo = ?, serial = ?, status = ?, motivo = ?,
                diretoria = ?, tipo = ?, marca = ?, processador = ?, memoria = ?,
                armazenamento = ?, possui_carregador = ?, recebido_por = ?,
                unidade = ?, observacoes = ?, movido_para_estoque = ?, email_responsavel = ?,
                gestor_email = ?, chamado_dell = ?
            WHERE id = ?
            """, (
                dados.get("usuario"), dados.get("nome"),
                dados.get("matricula"), dados.get("departamento"),
                dados.get("patrimonio"), dados.get("modelo"),
                dados.get("serial"), dados.get("status"),
                dados.get("motivo"), dados.get("diretoria"),
                dados.get("tipo"), dados.get("marca"),
                dados.get("processador"), dados.get("memoria"),
                dados.get("armazenamento"), dados.get("possui_carregador"),
                dados.get("recebido_por"), dados.get("unidade"),
                dados.get("observacoes"), dados.get("movido_para_estoque"),
                dados.get("email_responsavel"), dados.get("gestor_email"),
                dados.get("chamado_dell"), id_devolucao,
            ))
        conn.commit()

    return _sincronizar_planilha_seguro("após edição")


def excluir(id_devolucao):
    """Marca um registro como deletado (soft delete) — não apaga do banco."""
    agora = _agora_brasilia().strftime("%d/%m/%Y %H:%M")
    with conectar() as conn:
        conn.execute(
            "UPDATE devolucoes SET deletado_em = ? WHERE id = ?",
            (agora, id_devolucao),
        )
        conn.commit()

    return _sincronizar_planilha_seguro("após exclusão")


def atualizar_chamado_dell(id_devolucao, chamado):
    """Atualiza só o campo chamado_dell de um registro."""
    with conectar() as conn:
        conn.execute(
            "UPDATE devolucoes SET chamado_dell = ? WHERE id = ?",
            (chamado, id_devolucao),
        )
        conn.commit()

    return _sincronizar_planilha_seguro("após atualização de chamado Dell")


def registrar_email_enviado(id_devolucao, destino):
    """Marca no banco quando e para quem o email foi disparado."""
    agora = _agora_brasilia().strftime("%d/%m/%Y %H:%M")
    with conectar() as conn:
        conn.execute(
            "UPDATE devolucoes SET email_enviado_em = ? WHERE id = ?",
            (f"{agora} → {destino}", id_devolucao),
        )
        conn.commit()

    return _sincronizar_planilha_seguro("após registrar envio de email")


# ── Gestão de usuários ────────────────────────────────────────────

def criar_usuario(login, nome, senha, admin=False):
    """Cria um novo usuário. Lança ValueError se o login já existir."""
    login = login.strip().lower()
    if not login or not nome or not senha:
        raise ValueError("Login, nome e senha são obrigatórios.")
    agora = _agora_brasilia().strftime("%d/%m/%Y %H:%M")
    hash_senha = generate_password_hash(senha)
    with conectar() as conn:
        existente = conn.execute(
            "SELECT id FROM usuarios WHERE login = ?", (login,)
        ).fetchone()
        if existente:
            raise ValueError(f"Login '{login}' já existe.")
        conn.execute(
            "INSERT INTO usuarios (login, nome, senha_hash, admin, ativo, criado_em) VALUES (?, ?, ?, ?, 1, ?)",
            (login, nome.strip(), hash_senha, 1 if admin else 0, agora),
        )
        conn.commit()


def autenticar_usuario(login, senha):
    """Retorna o dict do usuário se credenciais corretas e ativo, senão None."""
    login = login.strip().lower()
    with conectar() as conn:
        row = conn.execute(
            "SELECT * FROM usuarios WHERE login = ? AND ativo = 1", (login,)
        ).fetchone()
    if row and check_password_hash(row["senha_hash"], senha):
        return _row_to_dict(row)
    return None


def listar_usuarios():
    """Retorna todos os usuários como lista de dicts."""
    with conectar() as conn:
        rows = conn.execute(
            "SELECT id, login, nome, admin, ativo, criado_em FROM usuarios ORDER BY admin DESC, nome"
        ).fetchall()
    return [_row_to_dict(r) for r in rows]


def desativar_usuario(id_usuario):
    """Desativa um usuário (não permite login)."""
    with conectar() as conn:
        conn.execute("UPDATE usuarios SET ativo = 0 WHERE id = ?", (id_usuario,))
        conn.commit()


def reativar_usuario(id_usuario):
    """Reativa um usuário desativado."""
    with conectar() as conn:
        conn.execute("UPDATE usuarios SET ativo = 1 WHERE id = ?", (id_usuario,))
        conn.commit()


def alterar_senha(id_usuario, nova_senha):
    """Atualiza a senha de um usuário."""
    with conectar() as conn:
        conn.execute(
            "UPDATE usuarios SET senha_hash = ? WHERE id = ?",
            (generate_password_hash(nova_senha), id_usuario),
        )
        conn.commit()


def usuario_existe(login):
    """Retorna True se já existe um usuário com esse login."""
    with conectar() as conn:
        row = conn.execute(
            "SELECT id FROM usuarios WHERE login = ?", (login.strip().lower(),)
        ).fetchone()
    return row is not None
