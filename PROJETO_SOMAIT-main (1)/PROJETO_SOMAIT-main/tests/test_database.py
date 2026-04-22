import sqlite3
import openpyxl
import pytest
import app.database as database


@pytest.fixture(autouse=True)
def temp_db(monkeypatch, tmp_path):
    db_file = tmp_path / "test_database.db"
    monkeypatch.setattr(database, "DB", str(db_file))
    database.criar()
    yield
    # cleanup is automatic with tmp_path


def test_inserir_e_listar():
    dados = {
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-0001",
        "modelo": "Teste",
        "serial": "SN0001",
        "status": "Devolvido",
        "motivo": "Teste",
        "foto": None,
    }

    database.inserir(dados)
    resultados = database.listar()

    assert len(resultados) == 1
    registro = resultados[0]
    assert registro["nome"] == "Test User"
    assert registro["serial"] == "SN0001"


def test_inserir_retorna_aviso_quando_sincronizacao_falha(monkeypatch):
    monkeypatch.setattr(database, "sincronizar_planilha_completa", lambda: (_ for _ in ()).throw(PermissionError("arquivo em uso")))

    registro = database.inserir({
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-LOCK",
        "modelo": "Teste",
        "serial": "SNLOCK",
        "status": "OK",
        "motivo": "Teste lock",
        "foto": None,
    })

    assert registro["patrimonio"] == "PT-LOCK"
    assert "planilha Excel está bloqueada" in registro["_sync_error"]


def test_alimenta_planilha_existente(monkeypatch, tmp_path):
    """O sistema cria as 3 abas automaticamente na planilha configurada."""
    planilha = tmp_path / "empresa.xlsx"
    # Arquivo não precisa existir — sistema cria automaticamente

    monkeypatch.setattr(database, "PLANILHA_EMPRESA", str(planilha))

    dados = {
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-0099",
        "modelo": "ModeloX",
        "serial": "SN0099",
        "status": "Devolvido",
        "motivo": "Teste planilha existente",
        "foto": None,
    }

    database.inserir(dados)

    wb2 = openpyxl.load_workbook(planilha)

    # 3 abas criadas automaticamente
    assert "Devoluções" in wb2.sheetnames
    assert "Resumo por Status" in wb2.sheetnames
    assert "Dell Danificados" in wb2.sheetnames

    ws = wb2["Devoluções"]
    # Linha 1 = título, linha 2 = cabeçalho, linha 3 = primeiro registro
    assert ws.max_row == 3
    # Coluna D = "LOGIN DE REDE" = usuario
    assert ws["D3"].value == "test.user"
    # Coluna K = "TAG" = patrimonio
    assert ws["K3"].value == "PT-0099"
    # Coluna G = "MODELO"
    assert ws["G3"].value == "ModeloX"


def test_alimenta_planilha_com_mapeamento_configuravel(monkeypatch, tmp_path):
    """Múltiplos registros são todos gravados; a aba Resumo conta corretamente."""
    planilha = tmp_path / "empresa_multi.xlsx"

    monkeypatch.setattr(database, "PLANILHA_EMPRESA", str(planilha))

    for i, status in enumerate(["OK", "Danificado", "OK"], 1):
        database.inserir({
            "usuario": f"user{i}",
            "nome": f"Colaborador {i}",
            "matricula": str(1000 + i),
            "departamento": "TI",
            "patrimonio": f"PT-{i:04d}",
            "modelo": "Modelo",
            "serial": f"SN{i:04d}",
            "status": status,
            "motivo": "Teste",
            "foto": None,
        })

    wb2 = openpyxl.load_workbook(planilha)

    ws_dev = wb2["Devoluções"]
    # Linha 1 = título, linha 2 = cabeçalho, linhas 3-5 = 3 registros
    assert ws_dev.max_row == 5

    ws_res = wb2["Resumo por Status"]
    # Verifica que "OK" aparece com contagem 2
    resumo = {ws_res.cell(row=r, column=1).value: ws_res.cell(row=r, column=2).value
              for r in range(2, ws_res.max_row + 1)}
    assert resumo.get("OK") == 2
    assert resumo.get("Danificado") == 1


def test_diagnostica_caminho_windows_incompativel_no_linux(monkeypatch):
    caminho_windows = r"C:\Compartilhado\TI\Devolucoes.xlsx"
    monkeypatch.setattr(database, "PLANILHA_EMPRESA", caminho_windows)

    diagnostico = database.diagnosticar_config_planilha()

    if database.os.name != "nt":
        assert diagnostico["ok"] is False
        assert diagnostico["bloqueada"] is True
        assert "Linux" in diagnostico["ambiente"]
    else:
        assert diagnostico["bloqueada"] is False


def test_nao_cria_arquivo_local_com_caminho_windows_no_linux(monkeypatch, tmp_path):
    caminho_windows = r"C:\Compartilhado\TI\Devolucoes.xlsx"
    monkeypatch.setattr(database, "PLANILHA_EMPRESA", caminho_windows)
    monkeypatch.chdir(tmp_path)

    dados = {
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-0200",
        "modelo": "ModeloX",
        "serial": "SN0200",
        "status": "OK",
        "motivo": "Teste bloqueio caminho",
        "foto": None,
    }

    database.inserir(dados)

    if database.os.name != "nt":
        assert not (tmp_path / caminho_windows).exists()


def _criar_registro(overrides=None):
    """Helper para criar um registro padrão."""
    dados = {
        "usuario": "test.user",
        "nome": "Test User",
        "matricula": "12345",
        "departamento": "TI",
        "patrimonio": "PT-0001",
        "modelo": "Teste",
        "serial": "SN0001",
        "status": "OK",
        "motivo": "Teste",
        "foto": None,
    }
    if overrides:
        dados.update(overrides)
    return database.inserir(dados)


def test_atualizar_registro():
    reg = _criar_registro({"patrimonio": "PT-UP01", "status": "OK"})
    database.atualizar(reg["id"], {
        "usuario": "new.user",
        "nome": "New Name",
        "matricula": "99999",
        "departamento": "RH",
        "patrimonio": "PT-UP01",
        "modelo": "NovoModelo",
        "serial": "SN-NEW",
        "status": "Danificado",
        "motivo": "Atualizado",
        "diretoria": "Corp",
        "tipo": "Desktop",
        "marca": "Lenovo",
        "processador": "i7",
        "memoria": "32 GB",
        "armazenamento": "SSD 512",
        "possui_carregador": "Não",
        "recebido_por": "Analista",
        "unidade": "SP",
        "observacoes": "Obs",
        "movido_para_estoque": "Sim",
        "email_responsavel": "e@e.com",
        "gestor_email": "g@e.com",
        "chamado_dell": "12345",
    })
    atualizado = database.buscar_por_id(reg["id"])
    assert atualizado["nome"] == "New Name"
    assert atualizado["status"] == "Danificado"
    assert atualizado["departamento"] == "RH"


def test_excluir_registro():
    reg = _criar_registro({"patrimonio": "PT-DEL"})
    assert database.buscar_por_id(reg["id"]) is not None
    database.excluir(reg["id"])
    assert database.buscar_por_id(reg["id"]) is None


def test_contar_filtrado():
    _criar_registro({"patrimonio": "PT-C1", "status": "OK"})
    _criar_registro({"patrimonio": "PT-C2", "status": "Danificado"})
    _criar_registro({"patrimonio": "PT-C3", "status": "OK"})
    assert database.contar_filtrado() == 3
    assert database.contar_filtrado(status="OK") == 2
    assert database.contar_filtrado(status="Danificado") == 1


def test_listar_filtrado_com_paginacao():
    for i in range(5):
        _criar_registro({"patrimonio": f"PT-P{i}", "status": "OK"})
    pagina1 = database.listar_filtrado(limit=2, offset=0)
    pagina2 = database.listar_filtrado(limit=2, offset=2)
    pagina3 = database.listar_filtrado(limit=2, offset=4)
    assert len(pagina1) == 2
    assert len(pagina2) == 2
    assert len(pagina3) == 1
    # Sem repetição entre páginas
    ids = [r["id"] for r in pagina1 + pagina2 + pagina3]
    assert len(ids) == len(set(ids))


def test_registrar_email_enviado():
    reg = _criar_registro({"patrimonio": "PT-EM01"})
    database.registrar_email_enviado(reg["id"], "RH")
    atualizado = database.buscar_por_id(reg["id"])
    assert "→ RH" in atualizado["email_enviado_em"]

