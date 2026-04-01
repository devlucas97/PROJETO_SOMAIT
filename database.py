import sqlite3
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
import os
 
DB = "database.db"
EXCEL_FILE = "devolucoes.xlsx"

def conectar():
    return sqlite3.connect(DB, detect_types=sqlite3.PARSE_DECLTYPES)
 
def criar():
    with conectar() as conn:
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS devolucoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT,
            usuario TEXT,
            nome TEXT,
            matricula TEXT,
            departamento TEXT,
            patrimonio TEXT,
            modelo TEXT,
            serial TEXT,
            status TEXT,
            motivo TEXT,
            foto TEXT
        )
        """)
        conn.commit()
 
def exportar_para_excel():
    """Gera/atualiza a planilha Excel com os dados atuais do banco"""
    registros = listar()

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Devolucoes" in wb.sheetnames:
            ws = wb["Devolucoes"]
            # Limpa a planilha, preservando estilo externo se houver
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Devolucoes")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devolucoes"

    headers = ["ID", "Data", "Usuario", "Nome", "Matricula", "Departamento", "Patrimonio", "Modelo", "Serial", "Status", "Motivo", "Foto"]
    ws.append(headers)

    for reg in registros:
        ws.append(list(reg))

    wb.save(EXCEL_FILE)


def inserir(dados):
    with conectar() as conn:
        cursor = conn.cursor()
        # Usa timezone de Brasília (America/Sao_Paulo)
        horario_brasilia = datetime.now(ZoneInfo("America/Sao_Paulo"))
        cursor.execute("""
        INSERT INTO devolucoes (
            data, usuario, nome, matricula, departamento,
            patrimonio, modelo, serial, status, motivo, foto
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            horario_brasilia.strftime("%d/%m/%Y %H:%M"),
            dados["usuario"],
            dados["nome"],
            dados["matricula"],
            dados["departamento"],
            dados["patrimonio"],
            dados["modelo"],
            dados["serial"],
            dados["status"],
            dados.get("motivo"),
            dados.get("foto")
        ))
        conn.commit()
    exportar_para_excel()
 
def inserir_com_data(dados):
    with conectar() as conn:
        cursor = conn.cursor()
        cursor.execute("""
        INSERT INTO devolucoes (
            data, usuario, nome, matricula, departamento,
            patrimonio, modelo, serial, status, motivo, foto
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            dados.get("data") or datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y %H:%M"),
            dados["usuario"],
            dados["nome"],
            dados["matricula"],
            dados["departamento"],
            dados["patrimonio"],
            dados["modelo"],
            dados["serial"],
            dados["status"],
            dados.get("motivo"),
            dados.get("foto")
        ))
        conn.commit()
 
def listar():
    with conectar() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM devolucoes ORDER BY id DESC")
        return cursor.fetchall()